#!/usr/bin/env python3
"""Export a detailed collectible workbook without writing Markdown reports."""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
import hashlib
import json
from pathlib import Path
import re
import sys
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(SCRIPT_DIR))
import audit_collectibles as audit  # noqa: E402

OUTPUT_PATH = PROJECT_ROOT / "收藏品全量详表.xlsx"
INVENTORY_CAPACITY = 20

RARITY_COLORS = {0: "F0E3C2", 1: "68D8FF", 2: "C987FF", 3: "FFD75A"}
RARITY_CARD_MARGINAL_PROBABILITIES = {0: 0.50, 1: 0.30, 2: 0.15, 3: 0.05}
BASE_DEFAULTS: dict[str, Any] = {
    "pickup_type": 0,
    "display_name": "移速道具",
    "drop_weight": 1.0,
    "description": "",
    "can_store_in_inventory": False,
    "stackable": False,
    "icon_scale": "Vector2(1, 1)",
    "heal_amount": 0,
    "duration": 5.0,
    "move_speed_multiplier": 1.0,
    "fire_rate_multiplier": 1.0,
    "player_form_mode": 0,
    "shot_pattern": 0,
    "requires_projectile_primary_attack": False,
    "requires_ammunition": False,
    "collectible_rarity": 0,
}

FIELD_LABELS = {
    "pickup_type": "拾取类型", "display_name": "显示名称", "drop_weight": "掉落权重",
    "description": "玩家可见描述", "can_store_in_inventory": "可存入背包", "stackable": "普通道具可叠加",
    "icon_texture": "图标资源引用", "icon_scale": "图标缩放", "duration": "普通效果时长",
    "collectible_effect_id": "收藏品效果ID", "collectible_rarity": "稀有度枚举",
    "collectible_stacks_by_copy": "逐份生效", "collectible_max_copies": "最大份数",
    "requires_projectile_primary_attack": "要求投射物普攻", "requires_ammunition": "要求弹药机制",
    "bullet_pierce_chance": "子弹穿透概率",
    "bullet_homing_chance": "子弹追踪概率", "ammo_free_shot_chance": "射击不耗弹概率",
    "collectible_ammo_capacity_additive_bonus": "弹匣容量加算值",
    "collectible_ammo_capacity_bonus_ratio": "弹匣容量提高比例",
    "collectible_reload_time_reduction": "换弹时间缩短比例",
    "skill_charge_preserve_chance": "技能不耗技力概率", "damage_against_burning_multiplier": "对燃烧目标伤害倍率",
    "damage_against_bleeding_multiplier": "对流血目标伤害倍率", "collectible_attack_bonus": "攻击力加值",
    "collectible_max_health_bonus": "生命上限加值", "collectible_move_speed_bonus": "移动速度加值",
    "collectible_attack_speed_bonus": "攻击速度加值",
    "collectible_dash_distance_bonus": "冲刺距离加值", "collectible_dash_cooldown_reduction": "冲刺冷却缩减",
    "collectible_physical_defense_bonus": "物理防御加值", "collectible_magic_defense_bonus": "魔法防御加值",
    "collectible_physical_damage_bonus": "物理伤害加值", "collectible_magic_damage_bonus": "法术伤害加值",
    "collectible_skill_charge_bonus_per_second": "每秒技力充能加值", "base_upgrade_free_chance": "基础升级免费概率",
    "incoming_ranged_front_damage_multiplier": "受到正面远程伤害倍率", "incoming_ranged_back_damage_multiplier": "受到背面远程伤害倍率",
    "incoming_ranged_dodge_chance": "远程攻击闪避概率", "attack_speed_xirang_step": "攻速息壤阶梯",
    "attack_speed_bonus_per_xirang_step": "每阶攻速加值", "defense_xirang_step": "双防息壤阶梯",
    "defense_bonus_per_xirang_step": "每阶双防加值", "periodic_effect_id": "周期效果ID",
    "periodic_interval": "周期间隔", "periodic_radius": "周期效果半径", "periodic_damage": "周期伤害",
    "periodic_attack_damage_multiplier": "周期攻击力倍率", "periodic_target_count": "周期目标数",
    "periodic_heal": "周期治疗量", "periodic_slow_multiplier": "周期减速后移速倍率",
    "periodic_slow_duration": "周期减速时长", "skill_effect_id": "技能触发效果ID",
    "skill_effect_radius": "技能效果半径", "skill_effect_duration": "技能效果时长",
    "skill_move_speed_multiplier": "技能触发移速倍率", "collectible_design_id": "独特设计ID",
    "collectible_design_note": "设计说明", "conditional_effect_id": "条件效果ID",
    "conditional_health_ratio_threshold": "条件生命比例阈值", "conditional_xirang_threshold": "条件息壤阈值",
    "conditional_attack_bonus": "条件攻击加值", "conditional_max_health_bonus": "条件生命上限加值",
    "conditional_move_speed_bonus": "条件移速加值", "conditional_physical_defense_bonus": "条件物防加值",
    "conditional_magic_defense_bonus": "条件法防加值", "conditional_physical_damage_bonus": "条件物伤加值",
    "conditional_magic_damage_bonus": "条件法伤加值", "conditional_skill_charge_bonus_per_second": "条件每秒技力加值",
    "conditional_bullet_pierce_chance": "条件穿透概率", "trigger_effect_id": "事件触发效果ID",
    "trigger_shot_interval": "每N次攻击触发", "trigger_cooldown": "触发冷却", "trigger_damage": "触发伤害",
    "trigger_radius": "触发半径", "trigger_heal": "触发治疗", "trigger_xirang": "触发息壤",
    "trigger_skill_charge": "触发技力", "trigger_slow_multiplier": "触发减速后移速倍率",
    "trigger_slow_duration": "触发减速时长", "on_hit_effect_id": "命中效果ID", "on_hit_chance": "命中触发概率",
    "on_hit_cooldown": "命中效果冷却", "on_hit_damage": "命中效果伤害", "on_hit_duration": "命中效果时长",
    "on_hit_tick_interval": "命中持续伤害间隔", "on_hit_radius": "命中效果半径", "on_hit_heal": "命中治疗",
    "on_hit_xirang": "命中息壤", "on_hit_skill_charge": "命中技力", "on_hit_slow_multiplier": "命中减速后移速倍率",
    "on_hit_physical_defense_modifier": "命中物防修正", "on_hit_damage_taken_multiplier": "目标受到伤害倍率",
    "on_hit_execute_health_ratio": "处决生命比例", "kill_effect_id": "击杀效果ID", "kill_cooldown": "击杀效果冷却",
    "kill_heal": "击杀治疗", "kill_xirang": "击杀息壤", "kill_skill_charge": "击杀技力", "kill_damage": "击杀伤害",
    "kill_radius": "击杀效果半径", "kill_duration": "击杀效果时长", "kill_slow_multiplier": "击杀减速后移速倍率",
    "kill_move_speed_multiplier": "击杀移速倍率", "_icon_res_path": "图标res路径",
    "script": "脚本ExtResource", "metadata/_custom_type_script": "自定义类型脚本UID",
}

FIELD_NOTES = {
    "collectible_max_copies": "仅逐份生效时有效；0表示配置不限制，但单个背包容量为20。",
    "collectible_stacks_by_copy": "关闭时，相同effect_id在运行时只生效一份，且商店不会再次提供。",
    "drop_weight": "收藏品商店不读取该字段；实际抽取先按稀有度权重，再在同稀有度内等概率选取。",
    "on_hit_damage_taken_multiplier": "1.22表示目标额外受到22%伤害；不是额外加22点伤害。",
    "incoming_ranged_front_damage_multiplier": "1.3表示受到的正面远程伤害提高30%。",
    "incoming_ranged_back_damage_multiplier": "0.5表示受到的背面远程伤害降低50%。",
    "damage_against_burning_multiplier": "1.2表示对燃烧敌人的伤害提高20%。",
    "damage_against_bleeding_multiplier": "1.2表示对流血敌人的伤害提高20%。",
    "periodic_slow_multiplier": "0.5表示目标剩余50%移动速度，即降低50%。",
    "on_hit_slow_multiplier": "0.78表示目标剩余78%移动速度，即降低22%。",
    "requires_projectile_primary_attack": "为true时，锄头猫等非投射物普攻角色不会在洛曦选项中看到该收藏品。",
    "requires_ammunition": "为true时，锄头猫等没有弹药/换弹机制的角色不会在洛曦选项中看到该收藏品。",
    "collectible_ammo_capacity_additive_bonus": "逐份求和；每种加算弹匣最多5份；在百分比容量乘算前生效。",
    "collectible_ammo_capacity_bonus_ratio": "同类只取背包中的最高值；最终容量=floor((基础容量+加算总和)×(1+最高比例))。",
    "collectible_reload_time_reduction": "同类只取背包中的最高值；有效换弹时间=基础换弹时间×(1-最高缩短比例)。",
}

RUNTIME_HANDLERS = {
    "基础数值": "Player._refresh_collectible_stats",
    "穿透": "Player._should_fire_piercing_bullet / Bullet",
    "追踪": "Player._should_fire_homing_bullet / Bullet",
    "弹药免耗": "Player普通射击耗弹判定",
    "弹匣容量加算": "Player._refresh_collectible_stats / AmmoRangedPlayer.get_ammo_capacity",
    "弹匣容量乘算": "Player._refresh_collectible_stats / AmmoRangedPlayer.get_ammo_capacity",
    "换弹缩短": "Player._refresh_collectible_stats / AmmoRangedPlayer.get_effective_reload_duration",
    "技能免耗": "Player技能发动与技力保留判定",
    "状态增伤": "Player.resolve_attack_damage_against_enemy",
    "免费升级": "Player/庄方宜基础升级消费判定",
    "远程防御": "Player.apply_damage方向与闪避判定",
    "息壤动态": "Player._refresh_collectible_stats",
    "周期": "Player._trigger_collectible_periodic_effect",
    "技能": "Player._activate_collectible_skill_effects",
    "条件": "Player._is_collectible_condition_active",
    "触发": "Player._apply_collectible_trigger_effect",
    "命中": "Player._apply_collectible_on_hit_effect",
    "击杀": "Player._apply_collectible_kill_effect",
}

WORDING_CHANGES = [
    ("先知方块", "攻击命中时，24%概率标记敌人3秒，使其受到伤害x1.22。", "攻击命中时，24%概率标记敌人3秒，使其额外受到22%伤害。"),
    ("咒刃", "攻击命中时，20%概率标记敌人2.5秒，使其受到伤害x1.18。", "攻击命中时，20%概率标记敌人2.5秒，使其额外受到18%伤害。"),
    ("盐晶符", "攻击命中时，12%概率标记敌人1.8秒，使其受到伤害x1.1。", "攻击命中时，12%概率标记敌人1.8秒，使其额外受到10%伤害。"),
    ("练习箭", "攻击命中时，12%概率标记敌人1.6秒，使其受到伤害x1.1。", "攻击命中时，12%概率标记敌人1.6秒，使其额外受到10%伤害。"),
    ("911", "正面远程伤害+30%；背面远程伤害-50%。", "受到的正面远程伤害提高30%；受到的背面远程伤害降低50%。"),
    ("银面具", "背面远程伤害-25%。", "受到的背面远程伤害降低25%。"),
]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def explicit_keys(path: Path) -> set[str]:
    result = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        if " = " in line and not line.startswith("["):
            result.add(line.split(" = ", 1)[0])
    return result


def effective_value(data: dict[str, Any], field: str) -> Any:
    if field in data:
        return data[field]
    if field in audit.FIELD_DEFAULTS:
        return audit.FIELD_DEFAULTS[field]
    return BASE_DEFAULTS.get(field, "")


def field_group(field: str) -> str:
    if field in {"pickup_type", "display_name", "drop_weight", "description", "can_store_in_inventory", "stackable", "duration"}:
        return "基础信息"
    if field.startswith("icon") or field == "_icon_res_path": return "显示资源"
    if field.startswith("periodic_"): return "周期效果"
    if field.startswith("skill_effect") or field == "skill_move_speed_multiplier": return "技能触发"
    if field.startswith("conditional_"): return "条件效果"
    if field.startswith("trigger_"): return "事件触发"
    if field.startswith("on_hit_"): return "命中效果"
    if field.startswith("kill_"): return "击杀效果"
    if field.startswith("collectible_design"): return "设计元数据"
    if field in {"script", "metadata/_custom_type_script"}: return "资源元数据"
    return "收藏品数值/规则"


def as_excel(value: Any) -> Any:
    if value is None: return ""
    if isinstance(value, (str, int, float, bool)): return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def style_sheet(ws, freeze="A2", filter_row=1) -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = ws.dimensions if filter_row else None
    for cell in ws[filter_row] if filter_row else []:
        cell.fill = PatternFill("solid", fgColor="26364A")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[filter_row].height = 32


def add_table(ws, name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1: return
    tab = Table(displayName=name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)


def autosize(ws, max_width=48, sample_rows=250) -> None:
    for col in range(1, ws.max_column + 1):
        values = [ws.cell(row, col).value for row in range(1, min(ws.max_row, sample_rows) + 1)]
        length = max((len(str(v)) for v in values if v is not None), default=8)
        ws.column_dimensions[get_column_letter(col)].width = min(max(length + 2, 10), max_width)


def write_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows: ws.append([as_excel(v) for v in row])


def specific_test_refs(config_name: str) -> list[str]:
    refs = []
    for path in sorted((PROJECT_ROOT / "dev_tools").glob("*smoke_test.gd")):
        text = path.read_text(encoding="utf-8")
        if config_name in text:
            refs.append(path.name)
    return refs


def main() -> None:
    audits = audit.audit_collectibles()
    if len(audits) != audit.EXPECTED_TOTAL_COUNT:
        raise RuntimeError(f"Expected {audit.EXPECTED_TOTAL_COUNT} collectibles, found {len(audits)}")
    if any(item.issues for item in audits):
        raise RuntimeError("Collectible audit failed: " + "; ".join(f"{a.path.name}: {a.issues}" for a in audits if a.issues))

    all_fields = sorted(set().union(*(item.data.keys() for item in audits)) | set(audit.FIELD_DEFAULTS.keys()))
    preferred = ["display_name", "description", "collectible_effect_id", "collectible_rarity", "collectible_stacks_by_copy", "collectible_max_copies"]
    all_fields = preferred + [f for f in all_fields if f not in preferred]
    explicit_by_name = {item.path.name: explicit_keys(item.path) for item in audits}

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.properties.title = "收藏品全量详表"
    wb.properties.subject = "Arc Nice 项目全部收藏品配置、机制、资源与审计信息"
    wb.properties.creator = "Codex"

    ws = wb.create_sheet("阅读指南")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1"); ws["A1"] = "收藏品全量详表"; ws["A1"].font = Font(size=22, bold=True, color="FFFFFF"); ws["A1"].fill = PatternFill("solid", fgColor="1F4E78"); ws["A1"].alignment = Alignment(horizontal="center")
    summary = [
        ("生成时间", datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %z")),
        ("正式收藏品总数", len(audits)), ("原有收藏品", sum(a.path.name in audit.ORIGINAL_CONFIG_NAMES for a in audits)),
        ("扩展收藏品", sum(a.path.name not in audit.ORIGINAL_CONFIG_NAMES for a in audits)), ("审计通过", sum(not a.issues for a in audits)),
        ("配置目录", "res://resources/config/collectibles"), ("背包容量", INVENTORY_CAPACITY),
        ("每轮洛曦可领取次数", 1), ("每次候选数", 3),
        ("每轮洛曦刷新", "最多4次，费用依次为100、200、500、1000息壤；下次休整期重置。"),
        ("抽取方式", "先抽整组三卡品质：普通×3占50%、稀有×3占30%、史诗×3占12%、2史诗+1传说占3%、1史诗+2传说占3%、传说×3占2%；再在目标品质池内无重复抽具体收藏品。"),
        ("工作簿口径", "有效值=配置显式值，否则使用PickupConfig/审计脚本默认值；底层倍率与自然语言换算同时保留。"),
    ]
    for i, (k, v) in enumerate(summary, 3): ws.cell(i, 1, k); ws.cell(i, 2, v)
    ws["D3"] = "工作表"; ws["E3"] = "内容"
    sheet_notes = [
        ("收藏品总表", "一行一件，适合筛选、策划审阅和横向比较。"), ("全参数矩阵", "所有配置字段的有效值，包含未显式写出的默认值。"),
        ("图标图鉴", "123件图标、名称、ID、稀有度的可视化总览。"), ("审计与资源", "配置/图标/导入文件状态、像素包围盒、文件大小、SHA-256。"),
        ("统计分析", "稀有度、机制类别、叠加规则、角色兼容性统计。"), ("系统与运行时", "获取、去重、叠加、背包、处理函数和联网实现说明。"),
        ("字段字典", "字段中文名、分组、默认值、单位/语义提示。"), ("测试与引用", "逐件的全量运行时覆盖与专项测试引用。"),
        ("文案改写记录", "本次倍率文案自然语言化的前后对照，底层数值未改。"),
    ]
    for i, (k, v) in enumerate(sheet_notes, 4): ws.cell(i, 4, k); ws.cell(i, 5, v)
    for cell in ("A3", "D3", "E3"): ws[cell].font = Font(bold=True, color="FFFFFF"); ws[cell].fill = PatternFill("solid", fgColor="26364A")
    ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 78; ws.column_dimensions["D"].width = 22; ws.column_dimensions["E"].width = 72
    for row in ws.iter_rows(min_row=3, max_row=20):
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws = wb.create_sheet("收藏品总表")
    headers = ["序号", "世代", "名称", "稀有度", "单卡边际概率", "效果ID", "玩家可见描述", "效果摘要", "主机制类别", "运行时处理入口", "逐份生效", "配置最大份数", "单背包实际最多份数", "重复获取规则", "要求投射物普攻", "要求弹药机制", "兼容锄头猫", "兼容维什戴尔", "兼容缇伊", "显式配置字段数", "非默认效果字段", "设计ID", "设计说明", "配置路径", "图标路径", "专项测试引用", "审计结论"]
    rows = []
    for index, item in enumerate(audits, 1):
        data = item.data; rarity = int(data["collectible_rarity"]); stacks = bool(data.get("collectible_stacks_by_copy", False)); cap = int(data.get("collectible_max_copies", 0))
        actual_cap = 1 if not stacks else min(cap, INVENTORY_CAPACITY) if cap > 0 else INVENTORY_CAPACITY
        categories = audit.primary_affix_categories(data) or (["特殊"] if data.get("collectible_effect_id") == "admin_doll" else [])
        handlers = sorted({RUNTIME_HANDLERS.get(c, "专用逻辑") for c in categories})
        refs = specific_test_refs(item.path.name)
        requires_projectile = bool(data.get("requires_projectile_primary_attack", False))
        requires_ammunition = bool(data.get("requires_ammunition", False))
        rows.append([index, "原有" if item.path.name in audit.ORIGINAL_CONFIG_NAMES else "扩展", data["display_name"], audit.RARITY_LABELS[rarity], RARITY_CARD_MARGINAL_PROBABILITIES[rarity], data["collectible_effect_id"], data["description"], item.effect_summary, "；".join(categories), "；".join(handlers), stacks, cap if stacks else 1, actual_cap, "可重复至上限" if stacks else "同效果ID仅一份", requires_projectile, requires_ammunition, not (requires_projectile or requires_ammunition), True, True, len(explicit_by_name[item.path.name]), json.dumps(dict(item.effect_signature), ensure_ascii=False), data.get("collectible_design_id", ""), data.get("collectible_design_note", ""), "res://" + item.path.relative_to(PROJECT_ROOT).as_posix(), "res://" + item.icon_path.relative_to(PROJECT_ROOT).as_posix(), "；".join(refs) or "无单件预载专项；由全量运行时审计覆盖", "通过" if not item.issues else "；".join(item.issues)])
    write_rows(ws, headers, rows); style_sheet(ws); add_table(ws, "CollectiblesSummary"); autosize(ws, 52)
    ws.column_dimensions["G"].width = 58; ws.column_dimensions["H"].width = 44; ws.column_dimensions["U"].width = 52; ws.column_dimensions["W"].width = 46
    for row in range(2, ws.max_row + 1): ws.cell(row, 5).number_format = "0%"
    for row in range(2, ws.max_row + 1):
        rarity = int(audits[row - 2].data["collectible_rarity"]); ws.cell(row, 4).fill = PatternFill("solid", fgColor=RARITY_COLORS[rarity])
        for col in [7, 8, 9, 10, 21, 23, 26, 27]: ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)

    ws = wb.create_sheet("全参数矩阵")
    matrix_headers = ["序号", "名称", "稀有度", "配置文件", "显式字段列表"] + [f"{FIELD_LABELS.get(f, f)}\n[{f}]" for f in all_fields]
    matrix_rows = []
    for index, item in enumerate(audits, 1):
        exp = explicit_by_name[item.path.name]
        matrix_rows.append([index, item.data["display_name"], audit.RARITY_LABELS[int(item.data["collectible_rarity"])], item.path.name, "；".join(sorted(exp))] + [effective_value(item.data, f) for f in all_fields])
    write_rows(ws, matrix_headers, matrix_rows); style_sheet(ws); add_table(ws, "AllEffectiveFields"); autosize(ws, 30)
    ws.column_dimensions["B"].width = 18; ws.column_dimensions["E"].width = 50
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws = wb.create_sheet("图标图鉴")
    ws.sheet_view.showGridLines = False
    columns_per_group = 4; groups_per_row = 4
    for index, item in enumerate(audits):
        group = index % groups_per_row; block_row = (index // groups_per_row) * 4 + 1; col = group * columns_per_group + 1
        rarity = int(item.data["collectible_rarity"])
        ws.merge_cells(start_row=block_row, start_column=col, end_row=block_row, end_column=col + 2)
        title = ws.cell(block_row, col, f"{index + 1:03d}  {item.data['display_name']}（{audit.RARITY_LABELS[rarity]}）")
        title.fill = PatternFill("solid", fgColor=RARITY_COLORS[rarity]); title.font = Font(bold=True); title.alignment = Alignment(horizontal="center")
        if item.icon_path and item.icon_path.is_file():
            image = ExcelImage(str(item.icon_path)); image.width = 64; image.height = 64; ws.add_image(image, ws.cell(block_row + 1, col).coordinate)
        ws.merge_cells(start_row=block_row + 1, start_column=col + 1, end_row=block_row + 1, end_column=col + 2)
        ws.cell(block_row + 1, col + 1, item.data["collectible_effect_id"]).alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=block_row + 2, start_column=col, end_row=block_row + 2, end_column=col + 2)
        ws.cell(block_row + 2, col, item.effect_summary).alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[block_row + 1].height = 52; ws.row_dimensions[block_row + 2].height = 34
    for col in range(1, groups_per_row * columns_per_group + 1): ws.column_dimensions[get_column_letter(col)].width = 10 if (col - 1) % 4 == 0 else 13

    ws = wb.create_sheet("审计与资源")
    headers = ["序号", "名称", "配置状态", "问题", "配置文件", "配置字节", "配置SHA-256", "图标文件", "图标字节", "图标SHA-256", "图标尺寸", "非透明包围盒", "主体宽", "主体高", "非透明像素数", "透明占比", "Godot导入文件", "导入状态", "配置最后修改", "图标最后修改"]
    rows = []
    for index, item in enumerate(audits, 1):
        icon = item.icon_path; bbox = item.icon_bbox; width = bbox[2] - bbox[0] if bbox else 0; height = bbox[3] - bbox[1] if bbox else 0; import_path = Path(str(icon) + ".import") if icon else None
        rows.append([index, item.data["display_name"], "通过" if not item.issues else "需处理", "；".join(item.issues), item.path.name, item.path.stat().st_size, sha256(item.path), icon.relative_to(PROJECT_ROOT).as_posix() if icon else "", icon.stat().st_size if icon and icon.is_file() else 0, sha256(icon) if icon and icon.is_file() else "", "32×32" if icon else "", str(bbox) if bbox else "", width, height, item.icon_alpha_pixels, 1.0 - item.icon_alpha_pixels / 1024.0, import_path.relative_to(PROJECT_ROOT).as_posix() if import_path else "", "存在" if import_path and import_path.is_file() else "缺失", datetime.fromtimestamp(item.path.stat().st_mtime).isoformat(sep=" ", timespec="seconds"), datetime.fromtimestamp(icon.stat().st_mtime).isoformat(sep=" ", timespec="seconds") if icon and icon.is_file() else ""])
    write_rows(ws, headers, rows); style_sheet(ws); add_table(ws, "AuditAndAssets"); autosize(ws, 46)
    ws.column_dimensions["G"].width = 68; ws.column_dimensions["J"].width = 68
    for row in range(2, ws.max_row + 1): ws.cell(row, 16).number_format = "0.00%"

    ws = wb.create_sheet("统计分析")
    ws.sheet_view.showGridLines = False
    rarity_counts = Counter(int(i.data["collectible_rarity"]) for i in audits)
    offer_paths = {
        0: "普通×3（50%）",
        1: "稀有×3（30%）",
        2: "史诗×3（12%）；与传说混合（6%）",
        3: "1传说（3%）；2传说（3%）；3传说（2%）",
    }
    write_rows(ws, ["稀有度", "数量", "占比", "单卡边际概率", "整组三卡出现方式"], [[audit.RARITY_LABELS[r], rarity_counts[r], rarity_counts[r] / len(audits), RARITY_CARD_MARGINAL_PROBABILITIES[r], offer_paths[r]] for r in range(4)])
    for row in range(2, 6): ws.cell(row, 3).number_format = ws.cell(row, 4).number_format = "0.00%"; ws.cell(row, 1).fill = PatternFill("solid", fgColor=RARITY_COLORS[row - 2])
    category_counts = Counter(c for i in audits for c in (audit.primary_affix_categories(i.data) or ["特殊"]))
    start = 8; ws.cell(start, 1, "机制类别"); ws.cell(start, 2, "涉及收藏品数")
    for offset, (cat, count) in enumerate(sorted(category_counts.items(), key=lambda x: (-x[1], x[0])), 1): ws.cell(start + offset, 1, cat); ws.cell(start + offset, 2, count)
    rule_counts = Counter("逐份叠加" if i.data.get("collectible_stacks_by_copy", False) else "唯一生效" for i in audits)
    ws.cell(1, 8, "叠加规则"); ws.cell(1, 9, "数量")
    for idx, pair in enumerate(rule_counts.items(), 2): ws.cell(idx, 8, pair[0]); ws.cell(idx, 9, pair[1])
    ws.cell(8, 8, "兼容性"); ws.cell(8, 9, "数量")
    projectile_count = sum(bool(i.data.get("requires_projectile_primary_attack", False)) for i in audits)
    ammunition_count = sum(bool(i.data.get("requires_ammunition", False)) for i in audits)
    hoe_cat_count = sum(not bool(i.data.get("requires_projectile_primary_attack", False) or i.data.get("requires_ammunition", False)) for i in audits)
    ws.cell(9, 8, "要求投射物普攻"); ws.cell(9, 9, projectile_count)
    ws.cell(10, 8, "要求弹药机制"); ws.cell(10, 9, ammunition_count)
    ws.cell(11, 8, "兼容锄头猫"); ws.cell(11, 9, hoe_cat_count)
    ws.cell(12, 8, "兼容维什戴尔/缇伊"); ws.cell(12, 9, len(audits))
    for cell in [ws["A1"], ws["B1"], ws["C1"], ws["D1"], ws["E1"], ws["A8"], ws["B8"], ws["H1"], ws["I1"], ws["H8"], ws["I8"]]: cell.fill = PatternFill("solid", fgColor="26364A"); cell.font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 20; ws.column_dimensions["E"].width = 32; ws.column_dimensions["H"].width = 28

    ws = wb.create_sheet("系统与运行时")
    rules = [
        ("配置发现", "LuoxiMerchant._get_collectible_config_paths", "扫描res://resources/config/collectibles下以collectible_开头的.tres并按路径排序。"),
        ("候选生成", "LuoxiMerchant._build_collectible_choices_from_pool", "每次先抽整组三卡品质模式，再按目标品质从兼容池内无放回抽3个具体收藏品；本地、刷新和多人主机共用。"),
        ("整组品质概率", "LuoxiMerchant.get_collectible_offer_rarity_pattern_for_roll", "普通×3为50%、稀有×3为30%、史诗×3为12%、2史诗+1传说为3%、1史诗+2传说为3%、传说×3为2%。"),
        ("每轮领取", "COLLECTIBLE_CLAIMS_PER_ROUND", "每名玩家每段场间时间最多成功领取1次；失败领取不消耗次数。"),
        ("付费刷新", "LuoxiMerchant.REFRESH_COSTS", "每段场间最多刷新4次，依次花费100、200、500、1000息壤；由单机游戏或多人主机权威扣费。"),
        ("刷新重置", "LuoxiMerchant.reset_intermission_state", "进入下一段休整期时同时清空领取计数、刷新计数和候选缓存。"),
        ("背包", "RunStateStore.INVENTORY_CAPACITY", "每名玩家20格；收藏品必须can_store_in_inventory=true。"),
        ("唯一效果", "LuoxiMerchant.is_collectible_available_for_inventory", "stacks_by_copy=false时，同effect_id已有一份就不会再次提供。"),
        ("叠加效果", "collectible_stacks_by_copy / collectible_max_copies", "逐份生效；max_copies=0代表配置不限，实际仍受20格背包限制。"),
        ("角色兼容", "Player.is_collectible_compatible", "requires_projectile_primary_attack和requires_ammunition分别过滤非投射物普攻、无弹药机制角色；本批弹药收藏品与橙子不会提供给锄头猫。"),
        ("弹匣容量", "AmmoRangedPlayer.get_ammo_capacity", "固定按floor((基础容量+所有加算容量)×(1+最高百分比容量加成))计算；百分比同类只取最高。"),
        ("换弹缩短", "AmmoRangedPlayer.get_effective_reload_duration", "有效换弹时间=基础换弹时间×(1-最高缩短比例)；同类只取最高。"),
        ("弹数迁移", "AmmoRangedPlayer._on_collectible_ammunition_stats_refreshed", "原本满弹时扩容同步补满；未满保持当前弹数；降容夹紧到新上限。"),
        ("全量刷新", "Player._refresh_collectible_stats", "从背包一次性聚合并缓存静态、条件、息壤动态、概率、弹药和防御倍率等效果。"),
        ("联网", "MpGame收藏品RPC/广播路径", "同步有效容量、当前弹数与换弹进度；有效容量不会覆盖角色基础ammo_capacity，避免重复乘算。"),
        ("通用运行时审计", "dev_tools/collectible_runtime_audit_smoke_test.gd", "遍历全部收藏品，验证字段、聚合属性、触发、命中、击杀、周期和技能效果。"),
        ("静态审计", "dev_tools/audit_collectibles.py", "检查123件数量、字段逻辑、唯一性、效果组合、图标尺寸/像素/导入文件和设计元数据。"),
    ]
    write_rows(ws, ["主题", "代码入口/常量", "详细说明"], rules); style_sheet(ws); add_table(ws, "RuntimeRules"); autosize(ws, 80)
    ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 52; ws.column_dimensions["C"].width = 88
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws = wb.create_sheet("字段字典")
    rows = []
    for field in all_fields:
        default = audit.FIELD_DEFAULTS.get(field, BASE_DEFAULTS.get(field, "资源文件显式提供/无通用默认"))
        explicit_count = sum(field in explicit_by_name[i.path.name] for i in audits)
        rows.append([field, FIELD_LABELS.get(field, field.replace("_", " ")), field_group(field), default, explicit_count, len(audits) - explicit_count, FIELD_NOTES.get(field, ""), "有效值列在“全参数矩阵”中；未显式写出时采用此默认值。"])
    write_rows(ws, ["字段名", "中文含义", "分组", "默认值", "显式配置件数", "使用默认值件数", "关键语义/单位", "读取提示"], rows); style_sheet(ws); add_table(ws, "FieldDictionary"); autosize(ws, 62)
    ws.column_dimensions["G"].width = 62; ws.column_dimensions["H"].width = 48

    ws = wb.create_sheet("测试与引用")
    rows = []
    for index, item in enumerate(audits, 1):
        refs = specific_test_refs(item.path.name)
        categories = audit.primary_affix_categories(item.data) or ["特殊"]
        scene_refs = ["scene/player/player.gd", "scene/luoxi_merchant.gd"]
        if any(c in categories for c in ["周期", "技能", "触发"]): scene_refs.append("scene/multiplayer/mp_game.gd")
        rows.append([index, item.data["display_name"], item.data["collectible_effect_id"], "是", "dev_tools/collectible_runtime_audit_smoke_test.gd", "；".join(refs) or "无逐件预载专项", "；".join(scene_refs), "；".join(categories)])
    write_rows(ws, ["序号", "名称", "效果ID", "全量运行时审计覆盖", "通用测试", "专项预载测试", "主要运行时文件", "覆盖机制"], rows); style_sheet(ws); add_table(ws, "TestReferences"); autosize(ws, 58)

    ws = wb.create_sheet("文案改写记录")
    rows = []
    by_name = {i.data["display_name"]: i for i in audits}
    for name, old, new in WORDING_CHANGES:
        item = by_name[name]; data = item.data
        raw_multiplier = data.get("on_hit_damage_taken_multiplier", data.get("incoming_ranged_front_damage_multiplier", data.get("incoming_ranged_back_damage_multiplier", "")))
        rows.append([name, old, new, raw_multiplier, "仅修改玩家可见描述；底层字段和值未修改", item.path.name])
    write_rows(ws, ["收藏品", "修改前", "修改后", "相关底层倍率", "数值影响", "配置文件"], rows); style_sheet(ws); add_table(ws, "WordingChanges"); autosize(ws, 76)
    ws.column_dimensions["B"].width = 68; ws.column_dimensions["C"].width = 68
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)

    thin = Side(style="thin", color="D9E2F3")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)
                if cell.alignment == Alignment(): cell.alignment = Alignment(vertical="top")

    wb.save(OUTPUT_PATH)
    check = load_workbook(OUTPUT_PATH, read_only=False, data_only=False)
    expected = {"收藏品总表": 124, "全参数矩阵": 124, "审计与资源": 124, "测试与引用": 124, "文案改写记录": 7}
    for name, rows in expected.items():
        if check[name].max_row != rows: raise RuntimeError(f"{name} row count mismatch: {check[name].max_row}")
    if len(check.sheetnames) != 10: raise RuntimeError(f"Expected 10 sheets, found {len(check.sheetnames)}")
    check.close()
    print(json.dumps({"output": str(OUTPUT_PATH), "collectibles": len(audits), "sheets": len(wb.sheetnames), "audit_issues": 0, "size": OUTPUT_PATH.stat().st_size}, ensure_ascii=False))


if __name__ == "__main__":
    main()
